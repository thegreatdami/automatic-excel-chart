from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

# 1- Lê pasta de trabalho e planilha
wb = load_workbook("data/pivot_table.xlsx")
sheet = wb["Relatório"]

# 2- Referências das linhas e colunas
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

# 3- Adicionando Dados e Categorias no Gráfico
barchat = BarChart()

data = Reference(
    sheet,
    min_col=min_column + 1,
    max_col=max_column,
    min_row=min_row,
    max_row=max_row
)

categories = Reference(
    sheet,
    min_col=min_column,
    max_col=min_column,
    min_row=min_row + 1,
    max_row=max_row
)

barchat.add_data(data, titles_from_data=True)
barchat.set_categories(categories)

# 4- Criando o Gráfico
sheet.add_chart(barchat, "B10")
barchat.title = "Vendas por Fabricantes"
barchat.style = 2

# 5- Salvando o Workbook
wb.save("data/barchat.xlsx")